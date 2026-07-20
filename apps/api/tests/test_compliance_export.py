import hashlib
import hmac
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.execution_session import ExecutionSession
from app.models.task import Task
from app.services.compliance_export import build_compliance_export_xlsx
from app.services.webhook_dispatcher import _sign_payload, dispatch_webhook_event


def test_build_compliance_export_xlsx_includes_checklist_columns(db: Session):
    task = Task(
        title="Export checklist task",
        description="Used for export test",
        outlet_id=1,
        assigned_to=1,
        created_by=1,
        priority="medium",
        status="completed",
    )
    db.add(task)
    db.flush()

    db.add(
        ExecutionSession(
            task_id=task.id,
            source_type="sop_task",
            status="completed",
            answers_json={
                "_checklist": {
                    "score": 72,
                    "status": "fail",
                    "failed_items": [
                        {"label": "Temperature log", "value": "12", "reason": "Out of range"}
                    ],
                }
            },
            submitted_by=1,
        )
    )
    db.commit()

    content = build_compliance_export_xlsx(db, all_outlets=True)
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    assert "Outlet" in headers
    assert "Score" in headers
    assert "Failed Items" in headers

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) >= 1
    row = rows[-1]
    assert row[1] == "Export checklist task"
    assert row[3] == 72
    assert row[4] == "Fail"
    assert "Temperature log" in str(row[5])


def test_dispatch_webhook_event_respects_enabled_flag(db: Session, monkeypatch):
    from app.models.app_settings import AppSettings
    import json as json_module

    db.add(
        AppSettings(
            key="workspace",
            payload=json_module.dumps({"webhook_enabled": False}),
        )
    )
    db.commit()

    called = {"value": False}

    def fake_post(*args, **kwargs):
        called["value"] = True

    monkeypatch.setattr("app.services.webhook_dispatcher._post_webhook", fake_post)

    delivered = dispatch_webhook_event(
        db,
        event_type="task.completed",
        payload={"task_id": 1},
        outlet_id=1,
    )

    assert delivered == 0
    assert called["value"] is False


def test_sign_payload_matches_hmac_sha256():
    body = b'{"event":"task.completed"}'
    signature = _sign_payload("test-secret", body)
    expected = hmac.new(b"test-secret", body, hashlib.sha256).hexdigest()
    assert signature == f"sha256={expected}"
